import requests
import pandas as pd
import openpyxl
import time
import urllib.parse
import os
import json
from datetime import datetime


def get_lowest_price(appid: int, market_hash_name: str, currency: int) -> float | None:
    """
    Fetch the lowest market price for a given Steam Market item.
    """
    url = "https://steamcommunity.com/market/priceoverview/"
    params = {"currency": currency, "appid": appid, "market_hash_name": market_hash_name}

    try:
        resp = requests.get(url, params=params, timeout=10)
        resp.raise_for_status()
        data = resp.json()
        if data.get("success") and data.get("lowest_price"):
            price_str = data["lowest_price"].replace("zł", "").replace(",", ".").strip()
            return float(price_str)
        print(f"❌ No price data for {market_hash_name}")
        return None
    except Exception as e:
        print(f"⚠️ Error fetching price for {market_hash_name}: {e}")
        return None


def ensure_link(item: str, appid: int) -> str:
    """
    Ensure the provided item name is converted into a valid Steam Market URL.
    """
    if item.startswith(("https://", "http://")):
        return item
    encoded = urllib.parse.quote(item)
    return f"https://steamcommunity.com/market/listings/{appid}/{encoded}"


def extract_name_from_link(item: str) -> str:
    """
    Extract the market hash name from a full Steam Market URL.
    """
    if "/730/" in item:
        try:
            part = item.split("/730/")[1]
            return urllib.parse.unquote(part)
        except Exception:
            return item
    return item


def create_excel_file_if_missing(file_path: str):
    """
    Create Excel file if it doesn't exist.
    If the folder doesn't exist, raise an error and stop execution.
    """
    folder = os.path.dirname(file_path)
    if folder and not os.path.exists(folder):
        raise FileNotFoundError(f"❌ Folder '{folder}' does not exist. Please create it first.")

    if not os.path.exists(file_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Init"
        ws.append(["This is an auto-generated file."])
        wb.save(file_path)
        print(f"🆕 Created new Excel file at: {file_path}")



def write_to_excel(file_path: str, rows: list[dict]) -> openpyxl.Workbook:
    """
    Write collected data into an Excel file.
    Creates the file if it does not exist, applies styling, colors, and formatting.
    """
    df = pd.DataFrame(rows)
    sheet_name = datetime.now().strftime("%Y-%m-%d_%H-%M")

    # Create a valid Excel file if it doesn't exist
    if not os.path.exists(file_path):
        folder = os.path.dirname(file_path)
        if folder and not os.path.exists(folder):
            raise FileNotFoundError(f"❌ Folder '{folder}' does not exist. Please create it first.")
        wb = openpyxl.Workbook()
        wb.save(file_path)
        print(f"🆕 Created new Excel file at: {file_path}")

    # Append new sheet with data
    with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="new") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)

    # Reopen workbook for formatting
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]

    # Center alignment and bold header styling
    alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    bold_font = openpyxl.styles.Font(bold=True)

    # Apply header formatting
    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.alignment = alignment
            cell.font = bold_font

    # Center all other cells
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = alignment

    # Highlight "% Return" column
    return_col = None
    for idx, col in enumerate(ws[1], 1):
        if col.value == "% Return":
            return_col = idx
            break

    if return_col:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=return_col)
            try:
                val = float(cell.value)
                if val >= 0:
                    # Green for profit
                    cell.fill = openpyxl.styles.PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                else:
                    # Red for loss
                    cell.fill = openpyxl.styles.PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            except (TypeError, ValueError):
                pass

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    wb.save(file_path)
    return wb




def generate_charts(wb: openpyxl.Workbook, file_path: str):
    """
    Generate line charts showing price trends for each tracked item.
    """
    import string

    for name in ["Charts", "ChartData"]:
        if name in wb.sheetnames:
            del wb[name]

    ws_data = wb.create_sheet("ChartData")
    ws_chart = wb.create_sheet("Charts")

    all_data = []
    first_sheet = wb[wb.sheetnames[0]]
    buy_prices = {row[1]: row[2] for row in first_sheet.iter_rows(min_row=2, values_only=True)}

    for sheet_name in wb.sheetnames:
        if sheet_name in ["Charts", "ChartData"]:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[3]:
                continue
            net_price = round(row[3] * 0.87, 2)
            all_data.append((sheet_name, row[1], net_price))

    if not all_data:
        print("No data for charts.")
        return

    df = pd.DataFrame(all_data, columns=["Date", "Name", "NetPrice"])
    pivot = df.groupby(["Date", "Name"], as_index=False)["NetPrice"].mean()
    pivot = pivot.pivot(index="Date", columns="Name", values="NetPrice").sort_index()

    ws_data.append(["Date"] + list(pivot.columns))
    for idx, row in pivot.iterrows():
        ws_data.append([idx] + list(row))

    charts_per_row = 2
    chart_width_cells = 15
    chart_height_rows = 29
    chart_row = 0
    chart_col = 0

    for col_idx in range(2, ws_data.max_column + 1):
        name = ws_data.cell(1, col_idx).value
        chart = openpyxl.chart.LineChart()
        chart.title = name
        chart.x_axis.title = "Date"
        chart.y_axis.title = "Net Price (after Steam fee)"
        chart.legend = None
        chart.height = 15
        chart.width = 25

        data_ref = openpyxl.chart.Reference(ws_data, min_col=col_idx, min_row=1, max_row=ws_data.max_row)
        cats_ref = openpyxl.chart.Reference(ws_data, min_col=1, min_row=2, max_row=ws_data.max_row)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        series = chart.series[0]
        series.graphicalProperties.line.width = 25000
        series.graphicalProperties.line.solidFill = "1F4E78"
        series.marker.symbol = "circle"
        series.marker.size = 6

        buy_price = buy_prices.get(name)
        if buy_price:
            extra_col = ws_data.max_column + 1
            ws_data.cell(1, extra_col, f"{name}_BuyPrice")
            for i in range(2, ws_data.max_row + 1):
                ws_data.cell(i, extra_col, buy_price)
            buy_ref = openpyxl.chart.Reference(ws_data, min_col=extra_col, min_row=1, max_row=ws_data.max_row)
            chart.add_data(buy_ref, titles_from_data=True)
            series_buy = chart.series[-1]
            series_buy.graphicalProperties.line.solidFill = "FF0000"
            series_buy.graphicalProperties.line.width = 40000
            series_buy.marker.symbol = "none"

        col_letter = string.ascii_uppercase[chart_col * chart_width_cells]
        row_pos = chart_row * chart_height_rows + 1
        ws_chart.add_chart(chart, f"{col_letter}{row_pos}")

        chart_col += 1
        if chart_col >= charts_per_row:
            chart_col = 0
            chart_row += 1

    wb._sheets.sort(key=lambda s: 0 if s.title == "Charts" else (1 if s.title == "ChartData" else 2))
    wb.save(file_path)


def fix_windows_path(path_str: str) -> str:
    """
    Return a clean absolute path using forward slashes.
    """
    normalized = os.path.abspath(os.path.expanduser(path_str))
    fixed = normalized.replace('\\', '/')
    return fixed


def load_config(config_path: str) -> dict:
    """
    Load configuration (items, currency, and output path) from JSON file.
    """
    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)


if __name__ == "__main__":
    config = load_config("config.json")
    appid = config["appid"]
    currency = config["currency"]
    file_out = config["output_file"]
    my_items = config["items"]
    
    save_path = fix_windows_path(file_out)

    create_excel_file_if_missing(save_path)

    rows = []
    for item, buy_price in my_items.items():
        link = ensure_link(item, appid)
        name = extract_name_from_link(link)

        price_now = get_lowest_price(appid, name, currency)
        if price_now is None:
            continue

        price_after_fee = price_now * 0.87
        profit_percent = ((price_after_fee - buy_price) / buy_price) * 100

        rows.append({
            "Item_Link": link,
            "Item_Name": name,
            "Buy_Price": buy_price,
            "Current_Sell_Price": round(price_now, 2),
            "Net_Sell_Price": round(price_after_fee, 2),
            "% Return": round(profit_percent, 2)
        })

        time.sleep(3)

    wb = write_to_excel(save_path, rows)
    generate_charts(wb, save_path)

    print(f"✅ Data saved and charts updated: {save_path}")
